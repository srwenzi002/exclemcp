from __future__ import annotations

import os
from pathlib import Path
from typing import Any

from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

mcp = FastMCP("excel-mcp")
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


def _workspace_root() -> Path:
    env_root = os.environ.get("EXCEL_MCP_ROOT")
    if env_root:
        return Path(env_root).expanduser().resolve()
    return Path.cwd().resolve()


def _safe_path(file_path: str) -> Path:
    candidate = Path(file_path).expanduser().resolve()
    workspace_root = _workspace_root()
    if workspace_root not in candidate.parents and candidate != workspace_root:
        raise ValueError(f"file_path must be inside workspace: {workspace_root}")
    if candidate.suffix.lower() not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Only {sorted(ALLOWED_EXTENSIONS)} files are supported")
    if candidate.exists() and candidate.is_dir():
        raise ValueError("file_path must point to a file, not a directory")
    return candidate


def _load_or_create_workbook(path: Path, create_if_missing: bool) -> Workbook:
    if path.exists():
        return load_workbook(path)
    if not create_if_missing:
        raise FileNotFoundError(f"Workbook not found: {path}")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.save(path)
    return wb


def _validate_sheet_name(sheet_name: str) -> None:
    if not sheet_name or len(sheet_name) > 31:
        raise ValueError("sheet_name must be 1-31 characters")
    invalid_chars = set(r'[]:*?/\\')
    if any(c in invalid_chars for c in sheet_name):
        raise ValueError("sheet_name contains invalid characters: []:*?/\\")


def _ensure_sheet(wb: Workbook, sheet_name: str, create_if_missing: bool) -> Worksheet:
    _validate_sheet_name(sheet_name)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    if not create_if_missing:
        raise ValueError(f"Sheet not found: {sheet_name}")
    return wb.create_sheet(title=sheet_name)


@mcp.tool()
def list_sheets(file_path: str, create_if_missing: bool = False) -> dict[str, Any]:
    """List all sheet names in an Excel workbook."""
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    return {"file_path": str(path), "sheets": wb.sheetnames, "workspace_root": str(_workspace_root())}


@mcp.tool()
def read_range(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    create_if_missing: bool = False,
) -> dict[str, Any]:
    """Read a range like A1:C10 and return values as a 2D array."""
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)

    rows = ws[cell_range]
    values = [[cell.value for cell in row] for row in rows]

    return {
        "file_path": str(path),
        "sheet_name": sheet_name,
        "cell_range": cell_range,
        "values": values,
    }


@mcp.tool()
def write_cell(
    file_path: str,
    sheet_name: str,
    cell: str,
    value: Any,
    create_if_missing: bool = True,
) -> dict[str, Any]:
    """Write one value into a single cell (for example B2)."""
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)

    ws[cell] = value
    wb.save(path)

    return {
        "file_path": str(path),
        "sheet_name": sheet_name,
        "cell": cell,
        "value": value,
        "saved": True,
    }


@mcp.tool()
def write_range(
    file_path: str,
    sheet_name: str,
    start_cell: str,
    values: list[list[Any]],
    create_if_missing: bool = True,
) -> dict[str, Any]:
    """Write a 2D array to sheet, starting at start_cell (for example A1)."""
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)

    start = ws[start_cell]
    start_row = start.row
    start_col = start.column

    written_cells = 0
    for r_idx, row_values in enumerate(values):
        for c_idx, v in enumerate(row_values):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=v)
            written_cells += 1

    wb.save(path)

    return {
        "file_path": str(path),
        "sheet_name": sheet_name,
        "start_cell": start_cell,
        "rows": len(values),
        "written_cells": written_cells,
        "saved": True,
    }


@mcp.tool()
def insert_rows(
    file_path: str,
    sheet_name: str,
    idx: int,
    amount: int = 1,
    create_if_missing: bool = False,
) -> dict[str, Any]:
    """Insert rows before idx (1-based)."""
    if idx < 1 or amount < 1:
        raise ValueError("idx and amount must be >= 1")
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)
    ws.insert_rows(idx=idx, amount=amount)
    wb.save(path)
    return {"file_path": str(path), "sheet_name": sheet_name, "idx": idx, "amount": amount, "saved": True}


@mcp.tool()
def delete_rows(
    file_path: str,
    sheet_name: str,
    idx: int,
    amount: int = 1,
    create_if_missing: bool = False,
) -> dict[str, Any]:
    """Delete rows from idx (1-based)."""
    if idx < 1 or amount < 1:
        raise ValueError("idx and amount must be >= 1")
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)
    ws.delete_rows(idx=idx, amount=amount)
    wb.save(path)
    return {"file_path": str(path), "sheet_name": sheet_name, "idx": idx, "amount": amount, "saved": True}


@mcp.tool()
def insert_columns(
    file_path: str,
    sheet_name: str,
    idx: int,
    amount: int = 1,
    create_if_missing: bool = False,
) -> dict[str, Any]:
    """Insert columns before idx (1-based)."""
    if idx < 1 or amount < 1:
        raise ValueError("idx and amount must be >= 1")
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)
    ws.insert_cols(idx=idx, amount=amount)
    wb.save(path)
    return {"file_path": str(path), "sheet_name": sheet_name, "idx": idx, "amount": amount, "saved": True}


@mcp.tool()
def delete_columns(
    file_path: str,
    sheet_name: str,
    idx: int,
    amount: int = 1,
    create_if_missing: bool = False,
) -> dict[str, Any]:
    """Delete columns from idx (1-based)."""
    if idx < 1 or amount < 1:
        raise ValueError("idx and amount must be >= 1")
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)
    ws.delete_cols(idx=idx, amount=amount)
    wb.save(path)
    return {"file_path": str(path), "sheet_name": sheet_name, "idx": idx, "amount": amount, "saved": True}


@mcp.tool()
def rename_sheet(
    file_path: str,
    old_name: str,
    new_name: str,
) -> dict[str, Any]:
    """Rename a worksheet."""
    _validate_sheet_name(new_name)
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=False)
    if old_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {old_name}")
    if new_name in wb.sheetnames:
        raise ValueError(f"Sheet already exists: {new_name}")
    wb[old_name].title = new_name
    wb.save(path)
    return {"file_path": str(path), "old_name": old_name, "new_name": new_name, "saved": True}


@mcp.tool()
def delete_sheet(file_path: str, sheet_name: str) -> dict[str, Any]:
    """Delete a worksheet (must leave at least one sheet)."""
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    if len(wb.sheetnames) == 1:
        raise ValueError("Cannot delete the only sheet in workbook")
    del wb[sheet_name]
    wb.save(path)
    return {"file_path": str(path), "deleted_sheet": sheet_name, "remaining_sheets": wb.sheetnames, "saved": True}


@mcp.tool()
def clear_range(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    create_if_missing: bool = False,
) -> dict[str, Any]:
    """Clear values in a range like A1:C10."""
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    cleared_cells = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.value = None
                cleared_cells += 1
    wb.save(path)
    return {
        "file_path": str(path),
        "sheet_name": sheet_name,
        "cell_range": cell_range,
        "cleared_cells": cleared_cells,
        "saved": True,
    }


@mcp.tool()
def format_range(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    bold: bool | None = None,
    wrap_text: bool | None = None,
    horizontal: str | None = None,
    vertical: str | None = None,
    number_format: str | None = None,
    fill_hex: str | None = None,
    create_if_missing: bool = False,
) -> dict[str, Any]:
    """Format cells in a range. fill_hex example: 'EAF2FF'."""
    path = _safe_path(file_path)
    wb = _load_or_create_workbook(path, create_if_missing=create_if_missing)
    ws = _ensure_sheet(wb, sheet_name, create_if_missing=create_if_missing)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    updated_cells = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if bold is not None:
                base_font = cell.font or Font()
                cell.font = Font(
                    name=base_font.name,
                    sz=base_font.sz,
                    italic=base_font.italic,
                    color=base_font.color,
                    underline=base_font.underline,
                    strike=base_font.strike,
                    bold=bold,
                )
            if any(v is not None for v in (wrap_text, horizontal, vertical)):
                base_alignment = cell.alignment or Alignment()
                cell.alignment = Alignment(
                    horizontal=horizontal if horizontal is not None else base_alignment.horizontal,
                    vertical=vertical if vertical is not None else base_alignment.vertical,
                    wrap_text=wrap_text if wrap_text is not None else base_alignment.wrap_text,
                    text_rotation=base_alignment.text_rotation,
                    shrink_to_fit=base_alignment.shrink_to_fit,
                    indent=base_alignment.indent,
                )
            if number_format is not None:
                cell.number_format = number_format
            if fill_hex is not None:
                color = fill_hex.strip().lstrip("#")
                if len(color) != 6:
                    raise ValueError("fill_hex must be 6 hex characters, e.g. EAF2FF")
                cell.fill = PatternFill(fill_type="solid", fgColor=color.upper())
            updated_cells += 1
    wb.save(path)
    return {
        "file_path": str(path),
        "sheet_name": sheet_name,
        "cell_range": cell_range,
        "updated_cells": updated_cells,
        "saved": True,
    }


def main() -> None:
    mcp.run()


if __name__ == "__main__":
    main()
